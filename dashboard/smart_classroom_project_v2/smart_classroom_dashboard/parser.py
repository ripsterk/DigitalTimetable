from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import tempfile
import uuid
from dataclasses import dataclass, asdict
from datetime import datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    import openpyxl
except ImportError as exc:
    raise RuntimeError("openpyxl is required. Run: pip install openpyxl") from exc

DAYS = {
    "LUNI": 0,
    "MARȚI": 1,
    "MARTI": 1,
    "MARŢI": 1,
    "MIERCURI": 2,
    "JOI": 3,
    "VINERI": 4,
    "SAMBATA": 5,
    "SÂMBĂTĂ": 5,
    "DUMINICA": 6,
    "DUMINICĂ": 6,
}

ROOM_RE = re.compile(r"\b([A-Z]{2})\s*[- ]?\s*(\d{2,3})\b", re.IGNORECASE)
PROF_RE = re.compile(
    r"\b((?:Prof\.?\s*(?:Dr\.?)?|Conf\.?(?:\s*Dr\.?)?|Ș\.?l\.?|S\.?l\.?|Lect\.?)\s*[^,;\-]*)",
    re.IGNORECASE,
)
TYPE_RE = re.compile(r"\((curs|lab|laborator|l|p|seminar|s|proiect|practic)\)", re.IGNORECASE)

@dataclass
class ScheduleEntry:
    id: str
    program: str
    sheet: str
    day: str
    day_index: int
    start: str
    end: str
    room: str
    room_key: str
    course: str
    professor: str
    activity_type: str
    group: str
    source_text: str
    source_cell: str


def normalize_day(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip().upper().replace("Ţ", "Ț").replace("Ş", "Ș")
    return s if s in DAYS else None


def normalize_room(room: str) -> str:
    m = ROOM_RE.search(room or "")
    if not m:
        return (room or "").replace(" ", "").upper()
    return f"{m.group(1).upper()} {int(m.group(2)):03d}"


def room_key(room: str) -> str:
    return normalize_room(room).replace(" ", "").upper()


def parse_hour_interval(value: Any) -> Optional[Tuple[int, int]]:
    if value is None:
        return None
    s = str(value).strip()
    m = re.search(r"(\d{1,2})\s*[-–]\s*(\d{1,2})", s)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def fmt_hour(h: int) -> str:
    return f"{h:02d}:00"


def get_merged_lookup(ws) -> Dict[Tuple[int, int], Any]:
    lookup = {}
    for r in ws.merged_cells.ranges:
        val = ws.cell(r.min_row, r.min_col).value
        for row in range(r.min_row, r.max_row + 1):
            for col in range(r.min_col, r.max_col + 1):
                lookup[(row, col)] = val
    return lookup


def merged_range_for_cell(ws, row: int, col: int):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng
    return None


def cell_value(ws, row: int, col: int, merged_lookup: Dict[Tuple[int, int], Any]):
    v = ws.cell(row, col).value
    if v is None:
        v = merged_lookup.get((row, col))
    return v


def find_header_row(ws) -> Optional[int]:
    for row in range(1, min(ws.max_row, 25) + 1):
        a = str(ws.cell(row, 1).value or "").strip().upper()
        b = str(ws.cell(row, 2).value or "").strip().upper()
        if "ZIUA" in a and "ORA" in b:
            return row
    return None


def clean_text(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip())


def extract_room(text: str) -> str:
    m = ROOM_RE.search(text)
    return normalize_room(m.group(0)) if m else "N/A"


def extract_professor(text: str) -> str:
    # Prefer explicit Prof./Conf./Ș.l.; otherwise return N/A.
    m = PROF_RE.search(text)
    if m:
        prof = clean_text(m.group(1)).strip(" -,")
        return prof if prof else "N/A"
    # Many entries store professor after final dash after the room.
    parts = [p.strip() for p in re.split(r"\s+-\s+", text) if p.strip()]
    if len(parts) >= 3:
        candidate = parts[-1]
        if not ROOM_RE.fullmatch(candidate):
            return candidate
    return "N/A"


def extract_activity_type(text: str) -> str:
    m = TYPE_RE.search(text)
    if not m:
        return "N/A"
    raw = m.group(1).lower()
    mapping = {"l": "laborator", "p": "practic", "s": "seminar", "lab": "laborator"}
    return mapping.get(raw, raw)


def extract_course(text: str) -> str:
    s = clean_text(text)
    # Remove professor segment first.
    s = re.sub(r",?\s*Prof\.?\s*[^,;\-]*", "", s, flags=re.I)
    s = re.sub(r",?\s*Conf\.?\s*[^,;\-]*", "", s, flags=re.I)
    s = re.sub(r"\s+-\s*Ș\.?l\.?[^,;\-]*", "", s, flags=re.I)
    # Split at room occurrence.
    m = ROOM_RE.search(s)
    if m:
        s = s[:m.start()].strip(" -,")
    # Drop trailing type/code parentheses after course name, keep the actual name before (curs/l/p).
    t = TYPE_RE.search(s)
    if t:
        s = s[:t.start()].strip(" -,")
    # If very short acronym-like entries, keep them.
    s = re.sub(r"\s+", " ", s).strip(" -,")
    return s if s else "N/A"


def sheet_program_name(ws) -> str:
    for row in range(1, min(10, ws.max_row) + 1):
        v = clean_text(ws.cell(row, 1).value)
        if "Master" in v:
            return v
    return ws.title


def parse_workbook_xlsx(xlsx_path: str | Path) -> List[ScheduleEntry]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    entries: List[ScheduleEntry] = []

    for ws in wb.worksheets:
        header = find_header_row(ws)
        if not header:
            continue
        merged_lookup = get_merged_lookup(ws)
        program = sheet_program_name(ws)

        # Fill day per row from col A merged labels.
        row_day: Dict[int, str] = {}
        current_day = None
        for r in range(header + 2, ws.max_row + 1):
            d = normalize_day(cell_value(ws, r, 1, merged_lookup))
            if d:
                current_day = d
            if current_day:
                row_day[r] = current_day

        seen_ranges = set()
        for r in range(header + 2, ws.max_row + 1):
            day = row_day.get(r)
            interval = parse_hour_interval(cell_value(ws, r, 2, merged_lookup))
            if not day or not interval:
                continue

            for c in range(3, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                text = clean_text(v)
                if not text or text.lower() == "disciplina si sala":
                    continue
                rng = merged_range_for_cell(ws, r, c)
                if rng:
                    key = str(rng)
                    if key in seen_ranges:
                        continue
                    seen_ranges.add(key)
                    start_row, end_row = rng.min_row, rng.max_row
                    source_cell = str(rng)
                    start_interval = parse_hour_interval(cell_value(ws, start_row, 2, merged_lookup))
                    end_interval = parse_hour_interval(cell_value(ws, end_row, 2, merged_lookup))
                    if not start_interval or not end_interval:
                        continue
                    start_h = start_interval[0]
                    end_h = end_interval[1]
                    group_col = rng.min_col
                else:
                    source_cell = ws.cell(r, c).coordinate
                    start_h, end_h = interval
                    group_col = c

                group = clean_text(cell_value(ws, header, group_col, merged_lookup))
                room = extract_room(text)
                entry = ScheduleEntry(
                    id=str(uuid.uuid4())[:8],
                    program=program,
                    sheet=ws.title,
                    day=day,
                    day_index=DAYS[day],
                    start=fmt_hour(start_h),
                    end=fmt_hour(end_h),
                    room=room,
                    room_key=room_key(room),
                    course=extract_course(text),
                    professor=extract_professor(text),
                    activity_type=extract_activity_type(text),
                    group=group or "N/A",
                    source_text=text,
                    source_cell=source_cell,
                )
                entries.append(entry)

    # Repair visual blocks where professor is stored in a separate row/cell.
    entries = repair_split_professor_rows(entries)

    # Sort for stable UI/API output.
    entries.sort(key=lambda e: (e.day_index, e.start, e.room_key, e.sheet, e.group))
    return entries


def convert_xls_to_xlsx(input_path: str | Path, output_dir: str | Path) -> Path:
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if input_path.suffix.lower() == ".xlsx":
        return input_path

    # Prefer LibreOffice when available, because it preserves merged cells well.
    soffice = shutil.which("libreoffice") or shutil.which("soffice")
    if soffice:
        cmd = [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(output_dir), str(input_path)]
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        converted = output_dir / (input_path.stem + ".xlsx")
        if converted.exists():
            return converted

    raise RuntimeError(
        "Could not convert .xls to .xlsx. Install LibreOffice on the server, or upload .xlsx instead."
    )


def parse_any_excel(input_path: str | Path) -> List[ScheduleEntry]:
    input_path = Path(input_path)
    with tempfile.TemporaryDirectory() as td:
        xlsx = convert_xls_to_xlsx(input_path, td)
        return parse_workbook_xlsx(xlsx)


def entries_to_json(entries: List[ScheduleEntry]) -> List[Dict[str, Any]]:
    return [asdict(e) for e in entries]


def save_entries(entries: List[ScheduleEntry], path: str | Path) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_text(json.dumps(entries_to_json(entries), ensure_ascii=False, indent=2), encoding="utf-8")


def load_entries(path: str | Path) -> List[Dict[str, Any]]:
    path = Path(path)
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8"))



def _is_na(value: Any) -> bool:
    s = clean_text(value)
    return not s or s.upper() in {"N/A", "NA", "NULL", "NONE", "-"}


def _time_to_minutes(value: str) -> int:
    h, m = str(value).split(":")[:2]
    return int(h) * 60 + int(m)


def _intervals_overlap_or_touch(a_start: str, a_end: str, b_start: str, b_end: str) -> bool:
    try:
        a1, a2 = _time_to_minutes(a_start), _time_to_minutes(a_end)
        b1, b2 = _time_to_minutes(b_start), _time_to_minutes(b_end)
        return max(a1, b1) <= min(a2, b2)
    except Exception:
        return False


def _entry_score(e: Dict[str, Any]) -> int:
    score = 0
    if not _is_na(e.get("course")):
        score += 100
    if not _is_na(e.get("professor")):
        score += 80
    if str(e.get("activity_type", "")).lower() == "curs":
        score += 15
    if str(e.get("sheet", "")).lower() == "manual":
        score += 20
    if str(e.get("group", "")).lower() in {"all", "manual"}:
        score += 5
    return score


def repair_split_professor_rows(entries: List[ScheduleEntry]) -> List[ScheduleEntry]:
    """Fix visual timetable rows where the course and professor are split.

    Some XLS schedules use one merged visual block for course information and a
    separate row/cell that contains only "EC 002, Prof. X". The old parser treated
    that as a separate entry with course=N/A, so the real course row kept
    professor=N/A. This pass copies that professor into the matching course row.
    """
    by_key: Dict[Tuple[str, int, str], List[ScheduleEntry]] = {}
    for e in entries:
        by_key.setdefault((e.sheet, e.day_index, e.room_key), []).append(e)

    for e in entries:
        if not _is_na(e.course) or _is_na(e.professor):
            continue

        # This is a professor-only row. Find nearby/overlapping course rows.
        candidates = by_key.get((e.sheet, e.day_index, e.room_key), [])
        best: Optional[ScheduleEntry] = None
        best_score = -1

        for c in candidates:
            if c is e:
                continue
            if _is_na(c.course):
                continue
            if not _is_na(c.professor):
                continue
            if not _intervals_overlap_or_touch(c.start, c.end, e.start, e.end):
                continue

            score = 0
            if c.group == e.group:
                score += 10
            if c.start <= e.start <= c.end:
                score += 5
            duration = abs(_time_to_minutes(c.start) - _time_to_minutes(e.start))
            score -= duration // 60

            if score > best_score:
                best_score = score
                best = c

        if best is not None:
            best.professor = e.professor
            best.source_text = f"{best.source_text} | professor row: {e.source_text}"

    return entries


def active_entry(entries: List[Dict[str, Any]], room: str, when: Optional[datetime] = None) -> Optional[Dict[str, Any]]:
    when = when or datetime.now()
    key = room_key(room)
    day_idx = when.weekday()
    t = when.strftime("%H:%M")
    candidates = []
    for e in entries:
        if e.get("room_key") != key:
            continue
        if e.get("day_index") != day_idx:
            continue
        if e.get("start") <= t < e.get("end"):
            candidates.append(e)

    if not candidates:
        return None

    # Prefer richer rows. This matters when the same visual timetable block
    # produces multiple rows, e.g. one row with the course and another with only
    # "EC 002, Prof. X".
    candidates.sort(key=_entry_score, reverse=True)
    return candidates[0]


def next_entry(entries: List[Dict[str, Any]], room: str, when: Optional[datetime] = None) -> Optional[Dict[str, Any]]:
    when = when or datetime.now()
    key = room_key(room)
    day_idx = when.weekday()
    t = when.strftime("%H:%M")
    todays = [e for e in entries if e.get("room_key") == key and e.get("day_index") == day_idx and e.get("start") > t]
    todays.sort(key=lambda e: e.get("start"))
    return todays[0] if todays else None

if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("file")
    p.add_argument("--out", default="data/schedule.json")
    args = p.parse_args()
    rows = parse_any_excel(args.file)
    save_entries(rows, args.out)
    print(f"Parsed {len(rows)} entries -> {args.out}")
    for row in rows[:10]:
        print(asdict(row))
